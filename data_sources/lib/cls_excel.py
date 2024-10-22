import openpyxl

class ExcelFile:
    def __init__(self, file_path, sheet=0):
        '''Init Excel object: book and sheet'''

        # Set default attributes
        self.workbook = None
        self.worksheet = None

        # Open an Excel file from `file_path`
        self.workbook = openpyxl.load_workbook(file_path)

        # Open an Excel sheet from `workbook`
        if isinstance(sheet, int):
            self.worksheet = self.workbook.worksheets[sheet]
        elif isinstance(sheet, str):
            self.worksheet = self.workbook[sheet]
        else:
            raise TypeError('Worksheet type error')


    def rows(self):
        '''Returns list of rows'''
        res = []
        for row in self.worksheet.iter_rows(values_only=True):
            ro = []
            for cell in row:
                ro.append(cell)
            res.append(ro)

        return res if res else False


    def columns(self):
        '''Returns list of columns'''
        res = []
        for col in self.worksheet.iter_cols():
            column = []
            for cell in col:
                column.append(cell.value)
            res.append(column)


        return res if res else False


    def get_dictionary(self) -> dict[str,list]:
        '''Returns dictionary from Excel file,
        first row as keys,
        other rows as list of values'''
        dictionary = {}
        for i in self.columns():
            dictionary.update({i[0]: i[1:]})
        return dictionary
